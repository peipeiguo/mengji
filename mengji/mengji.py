#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import sys
import datetime
import re
import logging
import docx
import win32com.client as win32
import configparser

from docx import Document
from openpyxl import load_workbook

VERSION = "1.0"
LOG_LEVEL = {'debug': logging.DEBUG, 'info': logging.INFO, 'warning': logging.WARNING, 'error': logging.error}

def get_config(section, key):
    path = os.path.dirname(os.path.abspath(__file__)) + '/mengji.cfg'
    if not os.path.exists(path):
        logging.error('Config file does not exist: "%s".', path)
        return None

    conf = configparser.ConfigParser()
    conf.read(path, encoding='utf-8')
    try:
        value = conf.get(section, key)
    except (configparser.NoSectionError, configparser.NoOptionError):
        value = None
    return value

# 将 .doc 文件转成 .docx 
def doc_to_docx(path):
    logging.info('Converting file "%s" to docx format...', path)
    word = win32.Dispatch('Word.Application') # 打开word应用程序
    word.Visible = 0        # 后台运行, 不显示
    word.DisplayAlerts = 0  # 不警告
    doc = word.Documents.Open(path) # 打开word文件
    new_path = os.path.splitext(path)[0] + '.docx'
    doc.SaveAs(new_path, 12, False, "", True, "", False, False, False, False)   # 12表示docx格式
    doc.Close()
    word.Quit()
    return new_path

def get_contract_files_in_directory(directory):
    '''Get contract file list in speicify directory
    '''
    logging.info('Getting contract files in "%s"...', directory)
    contract_files = []
    for file in os.listdir(directory):
        path = os.path.join(directory, file)
        ext = os.path.splitext(file)[1]
        if os.path.isfile(path) and (ext == ".doc" or ext == ".docx") and file.startswith("dp"):
            contract_files.append(path)

    logging.info('Total found %u contract files: %s', len(contract_files), contract_files)
    return contract_files

def read_contract_data_from_word(file_path, n):
    '''
    提取文件名 和文件名括号的内容 还有文档里表格里的 数据
    '''
    file_name = os.path.basename(file_path)
    file_name = os.path.splitext(file_name)[0]  # 去掉文件扩展名
    contract_no = re.sub(r'（[^）]+?）', '', file_name) # 合同号，提取自文件名
    customer = re.findall(r"（(.+?)）", file_name)[0]  # 客户名称，提取自文件名

    ext = os.path.splitext(file_path)[1]
    if ext == ".doc":
        file_path = doc_to_docx(file_path)
    
    logging.info('Reading contract data from file "%s"', file_path)
    doc = Document(file_path)
    table = doc.tables[n]
    orders = []  # 订单内容，提取自文档里表格里的数据
    # 忽略表格第一行（表头）和最后一行（表尾汇总信息）
    for i in range(1, len(table.rows) - 1):
        record = dict()
        record['subject'] = table.cell(i, 0).text.strip()       # 标的名称
        record['grade'] = table.cell(i, 1).text.strip()         # 牌号
        record['spec'] = table.cell(i, 2).text.strip()          # 规格型号
        record['quantity'] = table.cell(i, 4).text.strip()      # 数量
        record['unit'] = table.cell(i, 3).text.strip()          # 单位 
        record['unit_price'] = table.cell(i, 5).text.strip()    # 单价
        record['total_price'] = table.cell(i, 6).text.strip()   # 总价
        orders.append(record)

    if ext == ".doc":
        os.remove(file_path)
    return contract_no, customer, orders

def record_exists_in_excel(work_sheet, contract_no):
    '''Check if the record already exists
    '''
    for cell in work_sheet['C']:
        if cell.value == contract_no:
            return True
    return False

def append_contract_data_to_excel(file_path, sheet_name, contract_no, customer, orders):
    '''
    更新表格，主要更新合同号，客户名称，订单内容，数量，单位，单价，总价
    '''
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    #logging.info('表格：%s，工作表有：%s', file_path, wb.sheetnames) # 打印所有工作表的名称
    if sheet_name not in wb.sheetnames:
        logging.error('Sheet "%s" does not exists in file "%s".', sheet_name, file_path)
        return

    if record_exists_in_excel(ws, contract_no):
        logging.error('Data of contract No. "%s" already exists, please do not append data repeatedly.', contract_no)
        return

    logging.info('Start appending data to the worksheet "%s", total rows: %u, total columns: %u', sheet_name, ws.max_row, ws.max_column)
    for i in range(len(orders)):
        rec = dict()
        rec['C'] = contract_no      # 订单号（合同号）
        rec['D'] = customer         # 客户名称
        rec['E'] = orders[i]['subject'] + orders[i]['spec']   # 订单内容：标的名称+规格型号
        rec['J'] = float(orders[i]['quantity'])     # 开票数量
        rec['K'] = orders[i]['unit']         # 单位
        rec['L'] = float(orders[i]['unit_price'])   # 单价
        rec['M'] = float(orders[i]['total_price'])  # 结算金额
        rec['O'] = datetime.datetime.now().strftime("%Y/%m/%d") # 日期
        ws.append(rec)
        logging.info('Successfully appended one record: %s', rec)

    try:
        wb.save(file_path)
    except PermissionError: 
        logging.error('Can not save excel file, permission denied: "%s"', file_path)
    else:
        logging.info('Complete appending data to worksheet "%s"，total rows: %u, total columns: %u', sheet_name, ws.max_row, ws.max_column)
    wb.close()
    return

def summarize_contracts_to_account_form(contracts, account_form, sheet_name):
    '''Summarize all contract data to the specify account form
    '''
    if not len(contracts):
        logging.error('No contract documents, please check the path of contract documents.')
        return

    if not os.path.exists(account_form):
        logging.error('Account summary form file does not exist: "%s".', account_form)
        return

    for i in range(len(contracts)):
        logging.info('Starting to process %u/%u: contract: %s', i + 1, len(contracts), contracts[i])
        # 从合同文件中解析订单数据
        contract_no, customer, orders = read_contract_data_from_word(contracts[i], 0)
        logging.info('Contract no: %s, customer: %s', contract_no, customer)
        for i in range(len(orders)):
            logging.info('- Order content：%s, quantity: %s, unit: %s, unit price: %s, total price: %s', 
                orders[i]['subject'] + orders[i]['spec'],
                orders[i]['quantity'],
                orders[i]['unit'],
                orders[i]['unit_price'],
                orders[i]['total_price'])

        # 追加数据到账单汇总表格中info
        append_contract_data_to_excel(account_form, sheet_name, contract_no, customer, orders)
    return

if __name__ == "__main__":
    path = os.path.dirname(os.path.abspath(__file__)) + '\mengji.cfg'
    if not os.path.exists(path):
        print('ERROR: Config file does not exist: "%s".' % path)
        sys.exit()
    
    # 读取配置文件，设置日志信息
    log_file = get_config("General", "log-file")
    if log_file is None or len(log_file) == 0:
        log_file = 'mengji.log'

    level = get_config("General", "log-level")
    if level is None or len(level) == 0:
        level = 'info'
    logging.basicConfig(filename = log_file, format = '%(asctime)s %(filename)s:%(lineno)-3d %(levelname)-7s: %(message)s', level = LOG_LEVEL[level])
    logging.info('Starting Mengji %s, current working directory is "%s".', VERSION, os.getcwd())

    # 读取配置文件，设置合同目录
    doc_path = get_config("Source", "contract-doc-path")
    if doc_path is None or not os.path.exists(doc_path):
        logging.warning('Contract document path does not exsit: "%s", use current working directory instead.', doc_path)
        doc_path = os.getcwd()

    # 读取配置文件，设置汇总表单路径
    form_path = get_config("Destination", "account-form-path")
    sheet_name = get_config("Destination", "sheet-name")
    if form_path is None or len(form_path) == 0:
        logging.error('No account form path was specified in the configuration file.')
        logging.info('Exit Mengji %s.', VERSION)
        sys.exit()

    if sheet_name is None or len(sheet_name) == 0:
        logging.error('No sheet name was specified in the configuration file.')
        logging.info('Exit Mengji %s.', VERSION)
        sys.exit()

    # 从指定目录（默认：当前脚本文件所在目录）下获取合同文件列表
    contracts = get_contract_files_in_directory(doc_path)

    # 将合同文件中的订单数据汇总到指定的账单表格里（追加方式）
    logging.info('Destination info: sheet "%s" of account form "%s"', sheet_name, form_path)
    summarize_contracts_to_account_form(contracts, form_path, sheet_name)
    logging.info('Exit Mengji %s.', VERSION)
    sys.exit()

