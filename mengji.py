#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import re
import logging
import docx
import win32com.client as win32

from docx import Document
from openpyxl import load_workbook

# 将 .doc 文件转成 .docx 
def doc_to_docx(path):
    logging.info('Converting "%s" to docx format...', path)
    word = win32.Dispatch('Word.Application') # 打开word应用程序
    word.Visible = 0        # 后台运行, 不显示
    word.DisplayAlerts = 0  # 不警告
    doc = word.Documents.Open(path) # 打开word文件
    new_path = os.path.splitext(path)[0] + '.docx'
    doc.SaveAs(new_path, 12, False, "", True, "", False, False, False, False)   # 12表示docx格式
    doc.Close()
    word.Quit()
    return new_path

def get_contract_files_in_directory(directory):
    '''Get contract file list in speicify directory
    '''
    logging.info('Getting contract files in "%s"...', directory)
    contract_files = []
    for file in os.listdir(directory):
        if os.path.isfile(file) and os.path.splitext(file)[1] == ".doc" and file.startswith("dp"):
            logging.info(file)
            contract_files.append(os.path.join(directory, file))

    logging.info('Found %u files.', len(contract_files))
    return contract_files

def read_contract_data_from_word(file_path, n):
    '''
    提取文件名 和文件名括号的内容 还有文档里表格里的 数据
    '''
    file_name = os.path.basename(file_path)
    file_name = os.path.splitext(file_name)[0]  # 去掉文件扩展名
    contract_no = re.sub(r'（[^）]+?）', '', file_name) # 合同号，提取自文件名
    customer = re.findall(r"（(.+?)）", file_name)[0]  # 客户名称，提取自文件名

    file_path = doc_to_docx(file_path)
    logging.info('Reading contract data from "%s"', file_path)
    doc = Document(file_path)
    table = doc.tables[n]
    orders = []  # 订单内容，提取自文档里表格里的数据
    # 忽略表格第一行（表头）和最后一行（表尾汇总信息）
    for i in range(1, len(table.rows) - 1):
        record = dict()
        record['subject'] = table.cell(i, 0).text.strip()       # 标的名称
        record['grade'] = table.cell(i, 1).text.strip()         # 牌号
        record['spec'] = table.cell(i, 2).text.strip()          # 规格型号
        record['quantity'] = table.cell(i, 4).text.strip()      # 数量
        record['unit'] = table.cell(i, 3).text.strip()          # 单位 
        record['unit_price'] = table.cell(i, 5).text.strip()    # 单价
        record['total_price'] = table.cell(i, 6).text.strip()   # 总价
        orders.append(record)

    os.remove(file_path)
    return contract_no, customer, orders

def record_exists_in_excel(work_sheet, contract_no):
    '''Check if the record already exists
    '''
    for cell in work_sheet['C']:
        if cell.value == contract_no:
            return True
    return False

def append_contract_data_to_excel(file_path, sheet_name, contract_no, customer, orders):
    '''
    更新表格，主要更新合同号，客户名称，订单内容，数量，单位，单价，总价
    '''
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    logging.info('表格：%s，工作表有：%s', file_path, wb.sheetnames) # 打印所有工作表的名称

    if record_exists_in_excel(ws, contract_no):
        logging.info('合同号"%s"的数据已经存在，请勿重复汇总', contract_no)
        return

    logging.info('开始追加数据到工作表"%s"，总行数：%u，总列数：%u', sheet_name, ws.max_row, ws.max_column)
    for i in range(len(orders)):
        rec = dict()
        rec['C'] = contract_no      # 订单号（合同号）
        rec['D'] = customer         # 客户名称
        rec['E'] = orders[i]['subject'] + orders[i]['spec']   # 订单内容：标的名称+规格型号
        rec['J'] = orders[i]['quantity']     # 开票数量
        rec['K'] = orders[i]['unit']         # 单位
        rec['L'] = orders[i]['unit_price']   # 单价
        rec['M'] = orders[i]['total_price']  # 结算金额
        ws.append(rec)

    wb.save(file_path)
    logging.info('完成追加数据到工作表"%s"，总行数：%u，总列数：%u', sheet_name, ws.max_row, ws.max_column)
    wb.close()
    return

def summarize_contracts_to_account_form(contracts, account_form, sheet_name):
    '''Summarize all contract data to the specify account form
    '''
    if not len(contracts):
        logging.error('合同文件列表为空，请检查合同文件的存放位置是否正确。')
        return

    if not os.path.exists(account_form):
        logging.error('指定的账目汇总表单文件（"%s"）不存在。', account_form)
        return

    for i in range(len(contracts)):
        # 从合同文件中解析订单数据
        contract_no, customer, orders = read_contract_data_from_word(contracts[i], 0)
        logging.info('%u/%u: 合同号：%s，客户名称：%s', i + 1, len(contracts), contract_no, customer)
        for i in range(len(orders)):
            logging.info('订单内容：%s, 开票数量：%s，单位：%s，单价：%s，总价：%s', 
                orders[i]['subject'] + orders[i]['spec'],
                orders[i]['quantity'],
                orders[i]['unit'],
                orders[i]['unit_price'],
                orders[i]['total_price'])

        # 追加数据到账单汇总表格中info
        append_contract_data_to_excel(account_form, sheet_name, contract_no, customer, orders)
    return

if __name__ == "__main__":
    WORKING_DIR = os.getcwd()   # working directory
    ACCOUNT_FORM = "A1账目-大浦 (20-6-13).xlsx"
    SHEET_NAME = '订单2020'

    logging.basicConfig(format = '%(asctime)s %(levelname)s: %(message)s', level = logging.DEBUG)

    # 从指定目录（默认：当前脚本文件所在目录）下获取合同文件列表
    contracts = get_contract_files_in_directory(WORKING_DIR)

    # 将合同文件中的订单数据汇总到指定的账单表格里（追加方式）
    summarize_contracts_to_account_form(contracts, ACCOUNT_FORM, SHEET_NAME)


